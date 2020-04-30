from flask import Flask, flash, redirect, render_template, request, session, abort
import os
import json
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
app = Flask(__name__, static_url_path='/static')
db=json.load(open("database/db.json"))

@app.route('/')
def hrlogin():
    session['logged_in'] = False
    return render_template('HRlogin.html')

@app.route('/hrlogin', methods=["POST"])
def hlogin():
    username = request.form['Username']
    paswrd = request.form['Password']


    if (username in db['hrlogin']):
        if db['hrlogin'][username]['password']!=paswrd:
            return render_template('HRlogin.html', msg="Wrong password")
        else:
            session['logged_in'] = True
            return render_template('hrhhome.html')
    else:
        return render_template('HRlogin.html', msg="Wrong Username")
@app.route('/add_emp',methods=["POST"])
def addemp():
    id=request.form['empid']
    name = request.form['name']
    age = request.form['age']
    designation=request.form['designation']
    mn=request.form['mn']
    email = request.form['email']
    salary = request.form['salary']
    if id in db['employees_list']:
        return render_template('hrhhome.html',msg="Employee already exists")
    else:
        db['employees_list'].append(id)
        db['employees'][id]={
            "id":id,
            "name":name,
            "age":age,
            "designation":designation,
            "mobile":mn,
            "email":email,
            "salary":salary
        }
        json.dump(db, open("database/db.json", "w"))
        return render_template('hrhhome.html', msg="Employee added")

@app.route('/add_atn',methods=['POST'])
def add_attendance():
    id=request.form['empid']
    year = request.form['year']
    month = request.form['month']
    wd=request.form['present']
    if id not in db['employees_list']:
        return render_template('hrhhome.html', msg="Employee doesn't exists add employee first")
    else:
        wb=openpyxl.load_workbook('excel/attendance_employee.xlsx')
        ws=wb.active
        max_colum=ws.max_column
        flag=0
        max_row=ws.max_row
        for i in range(1,max_colum+1):
            cell=ws.cell(row=1,column=i)
            if cell.value==month:
                flag=1
                break
        if flag!=1:
            cell=ws.cell(row=1,column=i+1)
            cell.value=month
            flag=0
            i=i+1

        for j in range(1,max_row+1):
            cell=ws.cell(row=j,column=1)
            if cell.value==id:
                flag=1
                break
        if flag!=1:
            cell=ws.cell(row=j+1,column=1)
            cell.value=id
            j+=1
            flag=0
        cell=ws.cell(row=j,column=i)
        cell.value=wd
        wb.save("excel/attendance_employee.xlsx")
        return render_template('hrhhome.html', msg="Attendence added")

@app.route('/cal_sal',methods=['POST'])
def emp_verify():
    id = request.form['empid']
    month = request.form['month']
    if id not in db['employees_list']:
        return render_template('hrhhome.html', msg="Employee doesn't exists add employee first")
    else:
        if id in db['payslipgenerated'] and month in db['payslipgenerated'][id]['month']:
            return render_template('hrhhome.html', msg="Salary already generated")
        else:
            wb=openpyxl.load_workbook('excel/attendance_employee.xlsx')
            ws=wb.active
            max_colum=ws.max_column
            flag=0
            max_row=ws.max_row
            for i in range(1,max_colum+1):
                cell=ws.cell(row=1,column=i)
                if cell.value==month:
                    flag=1
                    break
            if flag!=1:
                return render_template('hrhhome.html', msg="Employee attendance is not entered add attendance for employee")
            for j in range(1,max_row+1):
                cell=ws.cell(row=j,column=1)
                if cell.value==id:
                    flag=1
                    break
            if flag!=1:
                return render_template('hrhhome.html', msg="Employee attendance for the month given is not entered")
            cell=ws.cell(row=j,column=i)
            pd=cell.value
            if pd == None:
                return render_template('hrhhome.html', msg="Employee attendance for the month given is not entered")
            else:
                payperday=db['employees'][id]['salary']
                basicpay=int(pd)*int(payperday)
                print(basicpay)
                return render_template('cal_sal.html', basicpay=basicpay,id=id,month=month)

@app.route('/cal_sal_emp',methods=['POST'])
def cal_Sal():
    id=request.form['empid']
    month = request.form['month']
    basicpay = request.form['basicpay']
    da=request.form['da']
    hra=request.form['hra']
    ta = request.form['ta']
    ca = request.form['ca']
    it = request.form['it']
    pt = request.form['pt']
    emi = request.form['emi']
    total_earnings=int(basicpay)+int(da)+int(hra)+int(ta)+int(ca)
    total_deductions=int(it)+int(pt)+int(emi)
    total_salary=total_earnings-total_deductions
    print(total_salary)
    f=open('payslip/payslip.txt',"w")
    f.write("ID:"+id+ os.linesep)
    f.write("Month:"+month+ os.linesep)
    f.write("Basicpay:"+ basicpay+ os.linesep)
    f.write("DA:"+ da+ os.linesep)
    f.write("HRA:"+hra+ os.linesep)
    f.write("CCA:"+ ca+ os.linesep)
    f.write("Transport allowance:"+ta+ os.linesep)
    f.write("Income tax:"+ it+ os.linesep)
    f.write("Proffesional tax:"+ pt+ os.linesep)
    f.write("EMI:"+ emi+ os.linesep)
    f.write("Total Earnings:"+str(total_earnings)+ os.linesep)
    f.write("Total deductuions:"+str(total_deductions)+ os.linesep)
    f.write("Total salary:"+str(total_salary)+ os.linesep)
    f.close()
    if id not in  db['payslipgenerated']:
        db['payslipgenerated'][id]={
            "month":[]
        }
    db['payslipgenerated'][id]['month'].append(month)
    json.dump(db, open("database/db.json", "w"))
    f=open("payslip/payslip.txt","r")
    message=f.read()
    f.close()
    sender = "cb.en.u4cse17541@cb.students.amrita.edu"
    password="Asdfgf;lkjhj"
    receivers = db['employees'][id]['email']

    s = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
    s.starttls()
    s.login(sender, password)
    msg = MIMEMultipart()
    print(message)
    msg['From'] = sender
    msg['To'] = receivers
    msg['Subject'] = "Payslip for "+month
    msg.attach(MIMEText(message, 'plain'))
    s.send_message(msg)
    s.quit()

    return render_template('hrhhome.html', msg="Pay slip sent to the employee")

@app.route('/mod_emp',methods=['POST'])
def mod_emp():
    id = request.form['empid']
    name = db['employees'][id]['name']
    age =  db['employees'][id]['age']
    designation =  db['employees'][id]['designation']
    mn =  db['employees'][id]['mobile']
    email =  db['employees'][id]['email']
    salary =  db['employees'][id]['salary']
    if request.form['btn1'] == 'Modify':
        return render_template('edit_employee.html',id=id,name=name,age=age,designation=designation,mobile=mn,email=email,salary=salary)
    elif request.form['btn1'] == 'Delete':
        if id in db['employees_list']:
           db['employees_list'].remove(id)
           del db['employees'][id]
           json.dump(db, open("database/db.json", "w"))
           return render_template('hrhhome.html', msg="Employee "+id +" is deleted")
        else:
            return render_template('hrhhome.html', msg="There is no such employee to delete")

@app.route('/edt_emp',methods=['POST'])
def edt_emp():
    id = request.form['empid']
    name = request.form['name']
    age = request.form['age']
    designation = request.form['designation']
    mn = request.form['mn']
    email = request.form['email']
    salary = request.form['salary']  
    db['employees'][id] = {
        "id": id,
        "name": name,
        "age": age,
        "designation": designation,
        "mobile": mn,
        "email": email,
        "salary": salary
    }
    json.dump(db, open("database/db.json", "w"))
    return render_template('hrhhome.html', msg="Employee " + id +" details updated")
@app.route('/back',methods=['GET', 'POST'])
def back():
    return render_template('hrhhome.html')
@app.route('/logout',methods=['GET', 'POST'])
def logout():
    session['logged-in']=False
    return render_template('HRlogin.html',msg="Successfully logged out")
@app.route('/hrhome',methods=['POST'])
def hr_home():
    if request.form['btn1']=='Add an Employee':
        return render_template('add_employee.html')
    elif request.form['btn1']=='Add Attendance':
        return render_template('add_attendance.html')
    elif request.form['btn1']=='Calculate Salary':
        return render_template('calculate_salary.html')
    elif request.form['btn1']=='Manage Employee Details':
        return render_template('manage_employee.html')




if __name__ == "__main__":
    app.secret_key = os.urandom(12)
    print("hello")
    app.run(debug=True, port=5000)